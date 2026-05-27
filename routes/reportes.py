from flask import Blueprint, render_template, request, send_file
from flask_login import login_required
from models import db, Venta, ItemVenta, Producto, Sede
from datetime import datetime, timedelta
from sqlalchemy import func
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pandas as pd

reportes_bp = Blueprint('reportes', __name__)

def _period_dates(periodo, meses=1):
    now = datetime.utcnow()
    periods = {
        'hoy': now.replace(hour=0, minute=0, second=0, microsecond=0),
        'semana': now - timedelta(days=7),
        'mes': now - timedelta(days=30),
        'trimestre': now - timedelta(days=90),
        'anual': now - timedelta(days=365),
        'custom': now - timedelta(days=30 * max(1, int(meses))),
    }
    return periods.get(periodo, periods['mes']), now

@reportes_bp.route('/reportes')
@login_required
def reportes():
    periodo  = request.args.get('periodo', 'mes')
    meses    = request.args.get('meses', '1')
    desde_str = request.args.get('desde', '')
    hasta_str = request.args.get('hasta', '')

    if desde_str and hasta_str:
        fecha_desde = datetime.strptime(desde_str, '%Y-%m-%d')
        fecha_hasta = datetime.strptime(hasta_str, '%Y-%m-%d') + timedelta(days=1)
    else:
        fecha_desde, fecha_hasta = _period_dates(periodo, meses)

    ventas = Venta.query.filter(
        Venta.fecha >= fecha_desde, Venta.fecha < fecha_hasta,
        Venta.estado == 'completada'
    ).all()

    total      = sum(v.total for v in ventas)
    total_costo = sum(iv.cantidad * (iv.producto.costo if iv.producto else 0)
                      for v in ventas for iv in v.items)
    ganancia   = total - total_costo
    margen     = (ganancia / total * 100) if total > 0 else 0
    ticket_avg = total / max(1, len(ventas))

    metodos = {}
    cats    = {}
    prod_v  = {}
    dias    = {}

    delta = min((fecha_hasta - fecha_desde).days, 90)
    for i in range(delta):
        d = (fecha_desde + timedelta(days=i)).strftime('%Y-%m-%d')
        dias[d] = 0

    for v in ventas:
        metodos[v.metodo_pago] = metodos.get(v.metodo_pago, 0) + v.total
        d = v.fecha.strftime('%Y-%m-%d')
        if d in dias:
            dias[d] += v.total
        for iv in v.items:
            cat = iv.producto.categoria.nombre if iv.producto and iv.producto.categoria else 'Sin categoría'
            cats[cat] = cats.get(cat, 0) + iv.subtotal
            k = iv.nombre_producto
            if k not in prod_v:
                prod_v[k] = {'qty': 0, 'total': 0, 'emoji': iv.producto.categoria.emoji if iv.producto and iv.producto.categoria else '📦'}
            prod_v[k]['qty']   += iv.cantidad
            prod_v[k]['total'] += iv.subtotal

    top_productos = sorted(prod_v.items(), key=lambda x: x[1]['total'], reverse=True)[:10]

    # Ventas por sede
    sedes_v = {}
    for v in ventas:
        s = v.sede.nombre if v.sede else 'Sin sede'
        sedes_v[s] = sedes_v.get(s, 0) + v.total

    # Métricas avanzadas
    clientes_unicos = len(set(v.cliente_nombre for v in ventas if v.cliente_nombre))
    total_unidades = sum(iv.cantidad for v in ventas for iv in v.items)
    conversion = (len(ventas) / max(1, total_unidades)) * 100
    ticket_promedio = total / max(1, len(ventas))
    roas = (total / max(1, total_costo)) if total_costo > 0 else 0
    margen_neto = margen
    ciclo_venta = 0  # Placeholder
    
    # Variación vs período anterior
    var_vs_anterior = 0
    proyeccion_venta = total * 1.15
    
    # Productos de bajo rendimiento
    low_performers = []
    for prod in Producto.query.filter_by(activo=True).all():
        ventas_prod = sum(iv.cantidad for v in ventas for iv in v.items if iv.producto_id == prod.id)
        if ventas_prod == 0 and prod.stock > 0:
            low_performers.append({'nombre': prod.nombre, 'ventas': 0})
        elif ventas_prod < 5 and prod.stock > 10:
            low_performers.append({'nombre': prod.nombre, 'ventas': ventas_prod})
    low_performers = low_performers[:5]

    return render_template('reportes.html',
        total=total, ganancia=ganancia, margen=margen,
        count=len(ventas), ticket_avg=ticket_avg,
        metodos=metodos, categorias_venta=cats,
        top_productos=top_productos,
        dias_labels=list(dias.keys()), dias_values=list(dias.values()),
        sedes_venta=sedes_v,
        periodo=periodo, meses=meses,
        fecha_desde=desde_str, fecha_hasta=hasta_str,
        ticket_promedio=ticket_promedio,
        clientes_unicos=clientes_unicos,
        total_unidades=total_unidades,
        conversion=conversion,
        roas=roas,
        margen_neto=margen_neto,
        var_vs_anterior=var_vs_anterior,
        proyeccion_venta=proyeccion_venta,
        low_performers=low_performers,
        ciclo_venta=ciclo_venta)

@reportes_bp.route('/exportar/excel')
@login_required
def exportar_excel():
    periodo = request.args.get('periodo', 'mes')
    meses = request.args.get('meses', '1')
    desde_str = request.args.get('desde', '')
    hasta_str = request.args.get('hasta', '')
    
    if desde_str and hasta_str:
        fecha_desde = datetime.strptime(desde_str, '%Y-%m-%d')
        fecha_hasta = datetime.strptime(hasta_str, '%Y-%m-%d') + timedelta(days=1)
    else:
        fecha_desde, fecha_hasta = _period_dates(periodo, meses)
    
    ventas = Venta.query.filter(
        Venta.fecha >= fecha_desde, Venta.fecha < fecha_hasta,
        Venta.estado == 'completada'
    ).all()
    
    total = sum(v.total for v in ventas)
    total_costo = sum(iv.cantidad * (iv.producto.costo if iv.producto else 0)
                      for v in ventas for iv in v.items)
    ganancia = total - total_costo
    margen = (ganancia / total * 100) if total > 0 else 0
    
    output = io.BytesIO()
    wb = Workbook()
    
    header_fill = PatternFill(start_color="D4A017", end_color="D4A017", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                    top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws_resumen = wb.active
    ws_resumen.title = "Resumen General"
    
    resumen_data = [
        ['MÉTRICA', 'VALOR'],
        ['Período', f'{fecha_desde.strftime("%d/%m/%Y")} - {fecha_hasta.strftime("%d/%m/%Y")}'],
        ['Total Ventas', f'${total:,.0f}'],
        ['N° Transacciones', len(ventas)],
        ['Ticket Promedio', f'${(total/len(ventas)):,.0f}' if ventas else '$0'],
        ['Ganancia Bruta', f'${ganancia:,.0f}'],
        ['Margen Bruto', f'{margen:.1f}%']
    ]
    
    for row, data_row in enumerate(resumen_data, 1):
        for col, value in enumerate(data_row, 1):
            cell = ws_resumen.cell(row=row, column=col, value=value)
            if row == 1:
                cell.fill = header_fill
                cell.font = header_font
            cell.border = border
    
    ws_resumen.column_dimensions['A'].width = 25
    ws_resumen.column_dimensions['B'].width = 30
    
    ws_ventas = wb.create_sheet("Ventas Detalladas")
    headers = ['Factura', 'Fecha', 'Cliente', 'Método Pago', 'Sede', 'Total', 'Items']
    for col, header in enumerate(headers, 1):
        cell = ws_ventas.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    row = 2
    for venta in ventas:
        items_count = sum(iv.cantidad for iv in venta.items)
        ws_ventas.cell(row=row, column=1, value=venta.numero_factura)
        ws_ventas.cell(row=row, column=2, value=venta.fecha.strftime('%Y-%m-%d %H:%M'))
        ws_ventas.cell(row=row, column=3, value=venta.cliente_nombre or 'Cliente general')
        ws_ventas.cell(row=row, column=4, value=venta.metodo_pago)
        ws_ventas.cell(row=row, column=5, value=venta.sede.nombre if venta.sede else 'Sin sede')
        ws_ventas.cell(row=row, column=6, value=f'${venta.total:,.0f}')
        ws_ventas.cell(row=row, column=7, value=items_count)
        row += 1
    
    wb.save(output)
    output.seek(0)
    
    nombre_archivo = f'reporte_ventas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=nombre_archivo
    )

@reportes_bp.route('/activos')
@login_required
def activos():
    productos = Producto.query.filter_by(activo=True).order_by(Producto.nombre).all()
    valor_costo = sum(p.costo * p.stock for p in productos)
    valor_venta = sum(p.precio_venta * p.stock for p in productos)

    sedes_r = {}
    resp_r  = {}
    for p in productos:
        s = p.sede.nombre if p.sede else 'Sin sede'
        r = p.responsable.nombre if p.responsable else 'Sin asignar'
        sedes_r.setdefault(s, {'count': 0, 'stock': 0, 'valor': 0})
        sedes_r[s]['count'] += 1
        sedes_r[s]['stock'] += p.stock
        sedes_r[s]['valor'] += p.precio_venta * p.stock
        resp_r.setdefault(r, {'count': 0})
        resp_r[r]['count'] += 1

    return render_template('activos.html',
        productos=productos, valor_costo=valor_costo, valor_venta=valor_venta,
        sedes_resumen=sedes_r, resp_resumen=resp_r)
