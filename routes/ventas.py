from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, session, send_file
from flask_login import login_required, current_user
from models import db, Venta, ItemVenta, Producto, Sede, Configuracion
from datetime import datetime, timedelta
from sqlalchemy import func, or_
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ventas_bp = Blueprint('ventas', __name__)

# ── helpers ────────────────────────────────────────────────────────────────

def _gen_factura():
    """Genera número de factura único"""
    last = Venta.query.order_by(Venta.id.desc()).first()
    num = (last.id + 1) if last else 1
    return f'D777-{num:06d}'

def _fmt(n):
    return f"${int(n):,}".replace(',', '.')

# ── POS ───────────────────────────────────────────────────────────────────

@ventas_bp.route('/pos')
@login_required
def pos():
    from models import Categoria
    categorias = Categoria.query.order_by(Categoria.nombre).all()
    sedes = Sede.query.filter_by(activa=True).all()
    # Limpiar carrito anterior
    session.pop('carrito', None)
    return render_template('pos.html', categorias=categorias, sedes=sedes)

# ── CARRITO (Session) ──────────────────────────────────────────────────────

@ventas_bp.route('/carrito/agregar', methods=['POST'])
@login_required
def carrito_agregar():
    """Agregar producto al carrito en sesión"""
    data = request.get_json()
    producto_id = data.get('producto_id')
    cantidad = data.get('cantidad', 1)
    
    producto = Producto.query.get(producto_id)
    if not producto or not producto.activo:
        return jsonify({'error': 'Producto no encontrado'}), 404
    
    if producto.stock < cantidad:
        return jsonify({'error': f'Stock insuficiente. Disponible: {producto.stock}'}), 400
    
    if 'carrito' not in session:
        session['carrito'] = []
    
    # Buscar si ya existe
    for item in session['carrito']:
        if item['producto_id'] == producto_id:
            nueva_cantidad = item['cantidad'] + cantidad
            if producto.stock < nueva_cantidad:
                return jsonify({'error': f'Stock insuficiente. Disponible: {producto.stock}'}), 400
            item['cantidad'] = nueva_cantidad
            item['subtotal'] = item['precio_unitario'] * item['cantidad']
            session.modified = True
            return jsonify({'success': True, 'carrito': session['carrito']})
    
    # Agregar nuevo
    session['carrito'].append({
        'producto_id': producto.id,
        'nombre': producto.nombre,
        'codigo_barras': producto.codigo_barras,
        'precio_unitario': producto.precio_venta,
        'cantidad': cantidad,
        'subtotal': producto.precio_venta * cantidad,
        'talla': producto.talla or '',
        'emoji': producto.categoria.emoji if producto.categoria else '📦'
    })
    session.modified = True
    
    return jsonify({'success': True, 'carrito': session['carrito']})

@ventas_bp.route('/carrito/eliminar', methods=['POST'])
@login_required
def carrito_eliminar():
    """Eliminar producto del carrito"""
    data = request.get_json()
    producto_id = data.get('producto_id')
    
    if 'carrito' in session:
        session['carrito'] = [item for item in session['carrito'] if item['producto_id'] != producto_id]
        session.modified = True
    
    return jsonify({'success': True, 'carrito': session.get('carrito', [])})

@ventas_bp.route('/carrito/actualizar', methods=['POST'])
@login_required
def carrito_actualizar():
    """Actualizar cantidad de producto en carrito"""
    data = request.get_json()
    producto_id = data.get('producto_id')
    cantidad = data.get('cantidad', 1)
    
    if cantidad <= 0:
        return carrito_eliminar()
    
    producto = Producto.query.get(producto_id)
    if not producto:
        return jsonify({'error': 'Producto no encontrado'}), 404
    
    if producto.stock < cantidad:
        return jsonify({'error': f'Stock insuficiente. Disponible: {producto.stock}'}), 400
    
    if 'carrito' in session:
        for item in session['carrito']:
            if item['producto_id'] == producto_id:
                item['cantidad'] = cantidad
                item['subtotal'] = item['precio_unitario'] * cantidad
                session.modified = True
                break
    
    return jsonify({'success': True, 'carrito': session.get('carrito', [])})

@ventas_bp.route('/carrito/limpiar', methods=['POST'])
@login_required
def carrito_limpiar():
    """Limpiar todo el carrito"""
    session.pop('carrito', None)
    return jsonify({'success': True})

@ventas_bp.route('/carrito/ver', methods=['GET'])
@login_required
def carrito_ver():
    """Ver contenido del carrito"""
    return jsonify({'carrito': session.get('carrito', [])})

# ── PROCESAR VENTA (JSON API) ──────────────────────────────────────────────

@ventas_bp.route('/ventas/procesar', methods=['POST'])
@login_required
def procesar_venta():
    data = request.get_json(silent=True)
    if not data:
        return jsonify({'ok': False, 'error': 'JSON inválido'}), 400

    # Usar carrito de sesión o items enviados
    if 'carrito' in session and not data.get('items'):
        items_data = session['carrito']
    else:
        items_data = data.get('items', [])
    
    if not items_data:
        return jsonify({'ok': False, 'error': 'Carrito vacío'}), 400

    # Validar stock primero
    productos = {}
    for it in items_data:
        pid = it.get('producto_id')
        cant = int(it.get('cantidad', 1))
        p = Producto.query.get(pid)
        if not p or not p.activo:
            return jsonify({'ok': False, 'error': f'Producto no encontrado: {pid}'}), 400
        if p.stock < cant:
            return jsonify({'ok': False, 'error': f'Stock insuficiente para {p.nombre} (disponible: {p.stock})'}), 400
        productos[pid] = (p, cant)

    subtotal = sum(it['precio_unitario'] * it['cantidad'] for it in items_data)
    desc_pct = max(0, min(100, float(data.get('descuento_pct', 0))))
    desc_monto = round(subtotal * desc_pct / 100, 2)
    total = round(subtotal - desc_monto, 2)
    sede_id = data.get('sede_id') or 1

    # Crear venta
    venta = Venta(
        numero_factura=_gen_factura(),
        fecha=datetime.utcnow(),
        cliente_nombre=data.get('cliente_nombre', 'Consumidor Final').strip() or 'Consumidor Final',
        cliente_doc=data.get('cliente_doc', '').strip(),
        metodo_pago=data.get('metodo_pago', 'efectivo'),
        subtotal=subtotal,
        descuento_pct=desc_pct,
        descuento_monto=desc_monto,
        total=total,
        notas=data.get('notas', '').strip(),
        estado='completada',
        usuario_id=current_user.id,
        sede_id=int(sede_id)
    )
    db.session.add(venta)
    db.session.flush()

    # Agregar items y actualizar stock
    for it in items_data:
        p, cant = productos[it['producto_id']]
        p.stock -= cant
        db.session.add(ItemVenta(
            venta_id=venta.id,
            producto_id=p.id,
            nombre_producto=p.nombre,
            codigo_barras=p.codigo_barras,
            cantidad=cant,
            precio_unitario=float(it['precio_unitario']),
            subtotal=float(it['precio_unitario']) * cant
        ))

    db.session.commit()
    
    # Limpiar carrito
    session.pop('carrito', None)
    
    return jsonify({
        'ok': True, 
        'venta_id': venta.id,
        'numero_factura': venta.numero_factura, 
        'total': total
    })

# ── HISTORIAL ─────────────────────────────────────────────────────────────

@ventas_bp.route('/ventas')
@login_required
def historial():
    desde = request.args.get('desde', '')
    hasta = request.args.get('hasta', '')
    metodo = request.args.get('metodo', '')
    sede_id = request.args.get('sede', '')

    q = Venta.query
    if desde:
        q = q.filter(Venta.fecha >= datetime.strptime(desde, '%Y-%m-%d'))
    if hasta:
        q = q.filter(Venta.fecha < datetime.strptime(hasta, '%Y-%m-%d') + timedelta(days=1))
    if metodo:
        q = q.filter_by(metodo_pago=metodo)
    if sede_id:
        q = q.filter_by(sede_id=int(sede_id))

    ventas = q.order_by(Venta.fecha.desc()).all()
    total_periodo = sum(v.total for v in ventas if v.estado == 'completada')
    total_items = sum(sum(i.cantidad for i in v.items) for v in ventas if v.estado == 'completada')
    ticket_avg = total_periodo / max(1, sum(1 for v in ventas if v.estado == 'completada'))
    sedes = Sede.query.filter_by(activa=True).all()

    return render_template('ventas.html',
        ventas=ventas, 
        total_periodo=total_periodo,
        total_items=total_items, 
        ticket_avg=ticket_avg,
        sedes=sedes, 
        fecha_desde=desde, 
        fecha_hasta=hasta,
        metodo_seleccionado=metodo,
        sede_seleccionada=sede_id)

# ── EXPORTAR A EXCEL ─────────────────────────────────────────────────────

@ventas_bp.route('/ventas/exportar/excel')
@login_required
def exportar_ventas_excel():
    """Exportar historial de ventas a Excel"""
    # Obtener los mismos parámetros del filtro
    desde = request.args.get('desde', '')
    hasta = request.args.get('hasta', '')
    metodo = request.args.get('metodo', '')
    sede_id = request.args.get('sede', '')
    
    # Aplicar mismos filtros que en historial
    q = Venta.query
    if desde:
        q = q.filter(Venta.fecha >= datetime.strptime(desde, '%Y-%m-%d'))
    if hasta:
        q = q.filter(Venta.fecha < datetime.strptime(hasta, '%Y-%m-%d') + timedelta(days=1))
    if metodo:
        q = q.filter_by(metodo_pago=metodo)
    if sede_id:
        q = q.filter_by(sede_id=int(sede_id))
    
    ventas = q.order_by(Venta.fecha.desc()).all()
    
    # Crear libro de Excel
    wb = Workbook()
    
    # Estilos
    header_fill = PatternFill(start_color="D4A017", end_color="D4A017", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                    top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Hoja de ventas
    ws = wb.active
    ws.title = "Historial de Ventas"
    
    # Encabezados
    headers = ['Factura', 'Fecha', 'Hora', 'Cliente', 'Documento', 'Método Pago', 
               'Subtotal', 'Descuento %', 'Descuento $', 'Total', 'Estado', 'Vendedor', 'Sede']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Agregar datos
    for row, venta in enumerate(ventas, 2):
        ws.cell(row=row, column=1, value=venta.numero_factura)
        ws.cell(row=row, column=2, value=venta.fecha.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=3, value=venta.fecha.strftime('%H:%M:%S'))
        ws.cell(row=row, column=4, value=venta.cliente_nombre or 'Consumidor Final')
        ws.cell(row=row, column=5, value=venta.cliente_doc or '')
        ws.cell(row=row, column=6, value=venta.metodo_pago or 'efectivo')
        ws.cell(row=row, column=7, value=venta.subtotal)
        ws.cell(row=row, column=8, value=venta.descuento_pct)
        ws.cell(row=row, column=9, value=venta.descuento_monto)
        ws.cell(row=row, column=10, value=venta.total)
        ws.cell(row=row, column=11, value=venta.estado)
        ws.cell(row=row, column=12, value=venta.usuario.nombre if venta.usuario else '')
        ws.cell(row=row, column=13, value=venta.sede.nombre if venta.sede else '')
        
        # Formato de moneda para columnas numéricas
        for col in [7, 9, 10]:
            cell = ws.cell(row=row, column=col)
            cell.number_format = '#,##0'
    
    # Ajustar anchos de columna
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Guardar archivo
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    nombre_archivo = f'ventas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=nombre_archivo
    )

# ── FACTURA ───────────────────────────────────────────────────────────────

@ventas_bp.route('/ventas/<int:vid>/factura')
@login_required
def ver_factura(vid):
    venta = Venta.query.get_or_404(vid)
    config = {
        k: Configuracion.get(f'tienda_{k}', d) for k, d in [
            ('nombre', 'DIFERENT 777'),
            ('nit', ''),
            ('direccion', ''),
            ('telefono', ''),
            ('email', ''),
            ('web', '')
        ]
    }
    return render_template('factura.html', venta=venta, config=config)

# ── ANULAR ────────────────────────────────────────────────────────────────

@ventas_bp.route('/ventas/<int:vid>/anular', methods=['POST'])
@login_required
def anular_venta(vid):
    if current_user.rol != 'admin':
        flash('Se requieren permisos de administrador', 'error')
        return redirect(url_for('ventas.historial'))
    
    v = Venta.query.get_or_404(vid)
    if v.estado == 'anulada':
        flash('Esta venta ya estaba anulada', 'info')
        return redirect(url_for('ventas.historial'))
    
    v.estado = 'anulada'
    for item in v.items:
        p = Producto.query.get(item.producto_id)
        if p:
            p.stock += item.cantidad
    
    db.session.commit()
    flash(f'Venta {v.numero_factura} anulada. Stock restaurado.', 'success')
    return redirect(url_for('ventas.historial'))

# ── BUSCAR PRODUCTO (scanner) ─────────────────────────────────────────────

@ventas_bp.route('/api/buscar-producto')
@login_required
def buscar_producto():
    codigo = request.args.get('codigo', '').strip()
    q = request.args.get('q', '').strip()
    
    if codigo:
        p = Producto.query.filter_by(codigo_barras=codigo, activo=True).first()
        if p and p.stock > 0:
            return jsonify({'found': True, 'producto': _prod_json(p)})
        return jsonify({'found': False, 'error': 'Sin stock' if p else 'No encontrado'})
    
    if q:
        ps = Producto.query.filter(
            Producto.activo == True, 
            Producto.stock > 0,
            or_(
                Producto.nombre.ilike(f'%{q}%'), 
                Producto.codigo_barras.ilike(f'%{q}%')
            )
        ).limit(12).all()
        return jsonify({'results': [_prod_json(p) for p in ps]})
    
    return jsonify({'results': []})

def _prod_json(p):
    return {
        'id': p.id, 
        'nombre': p.nombre, 
        'codigo_barras': p.codigo_barras,
        'precio_venta': p.precio_venta, 
        'stock': p.stock, 
        'talla': p.talla or '',
        'emoji': p.categoria.emoji if p.categoria else '📦',
        'categoria': p.categoria.nombre if p.categoria else '',
        'costo': p.costo
    }

# ── API PARA PRODUCTOS (POS) ───────────────────────────────────────────────

@ventas_bp.route('/api/productos')
@login_required
def api_productos():
    """API para obtener productos del POS"""
    sede_id = request.args.get('sede_id', type=int)
    query = Producto.query.filter_by(activo=True)
    
    if sede_id:
        query = query.filter_by(sede_id=sede_id)
    
    productos = query.order_by(Producto.nombre).all()
    
    return jsonify([{
        'id': p.id,
        'nombre': p.nombre,
        'codigo_barras': p.codigo_barras,
        'precio_venta': p.precio_venta,
        'stock': p.stock,
        'stock_minimo': p.stock_minimo,
        'talla': p.talla or '',
        'categoria': p.categoria.nombre if p.categoria else 'Sin categoría',
        'emoji': p.categoria.emoji if p.categoria else '📦',
        'activo': p.activo
    } for p in productos])

# ── ESTADÍSTICAS PARA DASHBOARD ────────────────────────────────────────────

@ventas_bp.route('/api/stats')
@login_required
def api_stats():
    """API para obtener estadísticas del dashboard"""
    hoy = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    semana = hoy - timedelta(days=7)
    mes = hoy - timedelta(days=30)
    
    # Ventas del día
    ventas_hoy = Venta.query.filter(
        Venta.fecha >= hoy,
        Venta.estado == 'completada'
    ).all()
    total_hoy = sum(v.total for v in ventas_hoy)
    count_hoy = len(ventas_hoy)
    
    # Ventas de la semana
    ventas_semana = Venta.query.filter(
        Venta.fecha >= semana,
        Venta.estado == 'completada'
    ).all()
    total_semana = sum(v.total for v in ventas_semana)
    
    # Ventas del mes
    ventas_mes = Venta.query.filter(
        Venta.fecha >= mes,
        Venta.estado == 'completada'
    ).all()
    total_mes = sum(v.total for v in ventas_mes)
    
    # Productos
    total_productos = Producto.query.filter_by(activo=True).count()
    stock_bajo = Producto.query.filter(
        Producto.activo == True,
        Producto.stock <= Producto.stock_minimo,
        Producto.stock > 0
    ).count()
    sin_stock = Producto.query.filter(
        Producto.activo == True,
        Producto.stock == 0
    ).count()
    
    # Datos para gráfico de últimos 7 días
    chart_labels = []
    chart_values = []
    for i in range(6, -1, -1):
        fecha = hoy - timedelta(days=i)
        fecha_inicio = fecha.replace(hour=0, minute=0, second=0)
        fecha_fin = fecha_inicio + timedelta(days=1)
        
        total_dia = db.session.query(func.sum(Venta.total)).filter(
            Venta.fecha >= fecha_inicio,
            Venta.fecha < fecha_fin,
            Venta.estado == 'completada'
        ).scalar() or 0
        
        chart_labels.append(fecha.strftime('%a'))
        chart_values.append(float(total_dia))
    
    # Ventas recientes
    ventas_recientes = Venta.query.filter_by(estado='completada')\
        .order_by(Venta.fecha.desc()).limit(10).all()
    
    # Top productos semana
    top_items = db.session.query(
        ItemVenta.nombre_producto,
        func.sum(ItemVenta.cantidad).label('total_cantidad'),
        func.sum(ItemVenta.subtotal).label('total_ventas')
    ).join(Venta).filter(
        Venta.fecha >= semana,
        Venta.estado == 'completada'
    ).group_by(ItemVenta.nombre_producto)\
     .order_by(func.sum(ItemVenta.subtotal).desc()).limit(10).all()
    
    return jsonify({
        'total_hoy': total_hoy,
        'count_hoy': count_hoy,
        'total_semana': total_semana,
        'total_mes': total_mes,
        'total_productos': total_productos,
        'stock_bajo': stock_bajo,
        'sin_stock': sin_stock,
        'chart_labels': chart_labels,
        'chart_values': chart_values,
        'ventas_recientes': [{
            'id': v.id,
            'numero_factura': v.numero_factura,
            'cliente_nombre': v.cliente_nombre,
            'fecha': v.fecha.strftime('%d/%m/%Y %H:%M'),
            'metodo_pago': v.metodo_pago,
            'total': v.total
        } for v in ventas_recientes],
        'top_items': [(item.nombre_producto, item.total_cantidad, float(item.total_ventas)) 
                      for item in top_items]
    })
